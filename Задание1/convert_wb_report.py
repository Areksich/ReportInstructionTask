#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт для преобразования отчетов Wildberries
Автоматически группирует данные по артикулам и рассчитывает показатели

Использование:
    python convert_wb_report.py input_file.xlsx
"""

import pandas as pd
import sys
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

if sys.platform.startswith('win'):
    # Пытаемся установить UTF-8 для консоли
    try:
        # Для Python 3.7+
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(encoding='utf-8')
            sys.stderr.reconfigure(encoding='utf-8')
        # Альтернативный способ для старых версий
        else:
            import codecs
            sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
            sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')
    except:
        # Если не получилось - продолжаем, но без эмодзи
        pass
    
    # Устанавливаем кодовую страницу UTF-8 для консоли Windows
    try:
        os.system('chcp 65001 > nul 2>&1')
    except:
        pass

# Безопасный print для Windows - если символ не поддерживается, заменяем на ASCII
def safe_print(text):
    """Безопасный вывод текста, работает даже если консоль не поддерживает UTF-8"""
    try:
        print(text)
    except UnicodeEncodeError:
        # Если не получилось - убираем проблемные символы
        ascii_text = text.encode('ascii', errors='ignore').decode('ascii')
        print(ascii_text)

# ============================================================================
# НАСТРОЙКИ (можешь менять)
# ============================================================================
COST_PERCENTAGE = 0.6  # Себестоимость = 60% от средней цены
SORT_BY = 'Продано (шт)'  # По какой колонке сортировать результат
SORT_ASCENDING = False  # False = от большего к меньшему

# ============================================================================
# ОСНОВНОЙ КОД
# ============================================================================

def load_report(file_path):
    """Загружает Excel файл с отчетом"""
    safe_print(f"[>] Загрузка файла: {file_path}")
    df = pd.read_excel(file_path)
    safe_print(f"[+] Загружено строк: {len(df)}")
    safe_print(f"[+] Уникальных артикулов: {df['Артикул поставщика'].nunique()}")
    return df

def process_report(df):
    """Обрабатывает данные и создает сводный отчет"""
    safe_print("\n[>] Обработка данных...")
    
    summary = []
    total_articles = df['Артикул поставщика'].nunique()
    processed = 0
    
    for article in df['Артикул поставщика'].unique():
        processed += 1
        if processed % 50 == 0:  # Показываем прогресс каждые 50 артикулов
            safe_print(f"    Обработано: {processed}/{total_articles} артикулов")
        
        article_data = df[df['Артикул поставщика'] == article]
        
        # Продажи
        sales_data = article_data[article_data['Тип документа'] == 'Продажа']
        qty_sold = sales_data['Кол-во'].sum()
        
        # Возвраты
        returns_data = article_data[article_data['Тип документа'] == 'Возврат']
        qty_returned = returns_data['Кол-во'].sum()
        
        # Средняя цена продажи (только по продажам)
        avg_price = sales_data['Цена розничная'].mean() if len(sales_data) > 0 else 0
        
        # Себестоимость = % от средней цены
        cost_price = avg_price * COST_PERCENTAGE
        total_cost = cost_price * qty_sold
        
        # Расходы
        logistics_cost = article_data['Услуги по доставке товара покупателю'].sum()
        penalties = article_data['Общая сумма штрафов'].sum()
        total_expenses = total_cost + logistics_cost + penalties
        
        # Доход
        revenue = article_data['К перечислению Продавцу за реализованный Товар'].sum()
        net_profit = revenue - total_expenses
        
        # Дополнительные данные
        product_name = article_data['Название'].iloc[0]
        brand = article_data['Бренд'].iloc[0] if 'Бренд' in article_data.columns else ''
        
        summary.append({
            'Артикул': article,
            'Бренд': brand,
            'Название товара': product_name,
            'Продано (шт)': int(qty_sold),
            'Возвращено (шт)': int(qty_returned),
            'Средняя цена продажи': round(avg_price, 2),
            'Себестоимость единицы': round(cost_price, 2),
            'Общая себестоимость': round(total_cost, 2),
            'Логистика': round(logistics_cost, 2),
            'Штрафы': round(penalties, 2),
            'Общие расходы': round(total_expenses, 2),
            'Выручка': round(revenue, 2),
            'Чистая прибыль': round(net_profit, 2),
            'Рентабельность, %': round((net_profit / revenue * 100) if revenue > 0 else 0, 2)
        })
    
    safe_print(f"[+] Обработано артикулов: {processed}/{total_articles}")
    
    # Создаем DataFrame и сортируем
    summary_df = pd.DataFrame(summary)
    summary_df = summary_df.sort_values(SORT_BY, ascending=SORT_ASCENDING).reset_index(drop=True)
    
    return summary_df

def add_total_row(df):
    """Добавляет итоговую строку"""
    total_row = {
        'Артикул': 'ИТОГО:',
        'Бренд': '',
        'Название товара': '',
        'Продано (шт)': df['Продано (шт)'].sum(),
        'Возвращено (шт)': df['Возвращено (шт)'].sum(),
        'Средняя цена продажи': '',
        'Себестоимость единицы': '',
        'Общая себестоимость': round(df['Общая себестоимость'].sum(), 2),
        'Логистика': round(df['Логистика'].sum(), 2),
        'Штрафы': round(df['Штрафы'].sum(), 2),
        'Общие расходы': round(df['Общие расходы'].sum(), 2),
        'Выручка': round(df['Выручка'].sum(), 2),
        'Чистая прибыль': round(df['Чистая прибыль'].sum(), 2),
        'Рентабельность, %': round((df['Чистая прибыль'].sum() / df['Выручка'].sum() * 100), 2)
    }
    
    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

def create_formatted_excel(df, output_path):
    """Создает красиво оформленный Excel файл"""
    safe_print("\n[>] Создание Excel файла...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводный отчет"
    
    # Заголовок
    ws['A1'] = 'СВОДНЫЙ ОТЧЕТ ПО ПРОДАЖАМ WILDBERRIES'
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:N1')
    ws.row_dimensions[1].height = 30
    
    # Описание
    ws['A2'] = f'Дата формирования: {datetime.now().strftime("%d.%m.%Y %H:%M")} | Себестоимость = {int(COST_PERCENTAGE*100)}% от средней цены'
    ws['A2'].font = Font(size=10, italic=True, color="666666")
    ws.merge_cells('A2:N2')
    ws.row_dimensions[2].height = 20
    
    # Заголовки колонок
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    ws.row_dimensions[3].height = 40
    
    # Данные
    for row_idx, row_data in enumerate(df.values, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Форматирование последней строки (ИТОГО)
            if row_idx == len(df) + 3:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            # Выравнивание
            if col_idx <= 3:  # Текстовые колонки
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Числовой формат
            if col_idx in [6, 7, 8, 9, 10, 11, 12, 13]:
                cell.number_format = '#,##0.00'
            elif col_idx == 14:  # Рентабельность %
                cell.number_format = '0.00"%"'
            
            # Цветовое выделение прибыли
            if col_idx == 13 and isinstance(value, (int, float)):
                if value < 0:
                    cell.font = Font(color="FF0000", bold=True)
                elif value > 0:
                    cell.font = Font(color="008000", bold=True)
            
            # Границы
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
    # Ширина колонок
    column_widths = {
        'A': 15, 'B': 15, 'C': 50, 'D': 12, 'E': 12,
        'F': 15, 'G': 15, 'H': 15, 'I': 12, 'J': 10,
        'K': 14, 'L': 14, 'M': 14, 'N': 14
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Закрепление заголовков
    ws.freeze_panes = 'A4'
    
    # Сохранение
    wb.save(output_path)
    safe_print(f"[+] Файл сохранен: {output_path}")

def print_summary(df):
    """Выводит краткую статистику"""
    total_row = df.iloc[-1]  # Последняя строка - итоги
    
    safe_print("\n" + "="*70)
    safe_print(">>> СТАТИСТИКА")
    safe_print("="*70)
    safe_print(f"Артикулов обработано:    {len(df) - 1}")
    safe_print(f"Всего продано:           {int(total_row['Продано (шт)'])} шт")
    safe_print(f"Всего возвращено:        {int(total_row['Возвращено (шт)'])} шт")
    safe_print(f"Процент возвратов:       {(total_row['Возвращено (шт)'] / total_row['Продано (шт)'] * 100):.2f}%")
    safe_print("")
    safe_print(f"Общая себестоимость:     {total_row['Общая себестоимость']:,.2f} руб")
    safe_print(f"Логистика:               {total_row['Логистика']:,.2f} руб")
    safe_print(f"Штрафы:                  {total_row['Штрафы']:,.2f} руб")
    safe_print(f"Общие расходы:           {total_row['Общие расходы']:,.2f} руб")
    safe_print("")
    safe_print(f"Выручка:                 {total_row['Выручка']:,.2f} руб")
    safe_print(f"Чистая прибыль:          {total_row['Чистая прибыль']:,.2f} руб")
    safe_print(f"Рентабельность:          {total_row['Рентабельность, %']:.2f}%")
    safe_print("="*70)

def main():
    """Основная функция"""
    safe_print("\n" + "="*70)
    safe_print(">>> ПРЕОБРАЗОВАНИЕ ОТЧЕТА WILDBERRIES")
    safe_print("="*70 + "\n")
    
    # Получаем путь к файлу
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    else:
        safe_print("Введите путь к файлу Excel:")
        input_file = input().strip().strip('"\'')
    
    input_path = Path(input_file)
    
    if not input_path.exists():
        safe_print(f"[!] Ошибка: Файл не найден: {input_path}")
        return
    
    # Создаем имя выходного файла
    output_path = input_path.parent / f"Сводный_отчет_{input_path.stem}_{datetime.now().strftime('%d.%m.%Y')}.xlsx"
    
    try:
        # Обработка
        df = load_report(input_path)
        summary_df = process_report(df)
        summary_df_with_total = add_total_row(summary_df)
        create_formatted_excel(summary_df_with_total, output_path)
        print_summary(summary_df_with_total)
        
        safe_print(f"\n[+] ГОТОВО! Файл сохранен в:")
        safe_print(f"    {output_path}")
        
    except Exception as e:
        safe_print(f"\n[!] Ошибка при обработке: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
